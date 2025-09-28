from __future__ import annotations
from typing import Dict, Any, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

CM_TO_POINTS = 72.0 / 2.54  # points per cm

def add_back_page_fn(
    wb: Workbook,
    class_no: Optional[str | int] = None,
    division: Optional[str] = None,
    subjects: Optional[List[Dict[str, Any]]] = None,
    catalog_doc: Optional[Dict[str, Any]] = None,
) -> Workbook:
    ws = wb.create_sheet(title="Back Page")

    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
    ws.page_margins.left = 0.0
    ws.page_margins.right = 0.0
    ws.page_margins.top = 0.3937
    ws.page_margins.bottom = 0.3937

    ws.column_dimensions['A'].width = 5.51
    for c in ['B','C','D','E','F','G','H','I','J','K']:
        ws.column_dimensions[c].width = 8.92
    ws.column_dimensions['L'].width = 13.33

    ws.row_dimensions[1].auto_height = False
    ws.row_dimensions[1].height = 0.91 * CM_TO_POINTS
    ws.row_dimensions[2].auto_height = False
    ws.row_dimensions[2].height = 0.85 * CM_TO_POINTS

    ws.merge_cells('B2:C2')
    ws.merge_cells('D2:H2')
    ws.merge_cells('I2:K2')

    font_name = "Kokila"
    hf = Font(name=font_name, size=14, bold=True)
    ws['B2'].value = "विषय"
    ws['D2'].value = "महिन्यात पूर्ण केलेला अभ्यासक्रम"
    ws['I2'].value = "विषय शिक्षकाची सही"
    for a in ['B2','D2','I2']:
        ws[a].alignment = Alignment(horizontal='center', vertical='center')
        ws[a].font = hf

    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    def thin_range(rng: str):
        for row in ws[rng]:
            for cell in row:
                cell.border = thin_border
    thin_range('B2:C2'); thin_range('D2:H2'); thin_range('I2:K2')

    def edge(r: int, c: int, left=None, right=None, top=None, bottom=None):
        e = ws.cell(r, c).border
        ws.cell(r, c).border = Border(
            left=left if left is not None else e.left,
            right=right if right is not None else e.right,
            top=top if top is not None else e.top,
            bottom=bottom if bottom is not None else e.bottom,
        )
    medium = Side(style="medium")

    for col in (2, 3):
        edge(2, col, top=medium); edge(2, col, bottom=medium)
    edge(2, 2, left=medium); edge(2, 3, right=medium)

    for col in range(4, 12):
        edge(2, col, top=medium); edge(2, col, bottom=medium)
    edge(2, 4, left=medium); edge(2, 11, right=medium)

    rows = [s for s in (subjects or []) if isinstance(s, dict) and s.get('active', True)]
    rows.sort(key=lambda s: s.get('order', 0))

    r = 3
    for s in rows:
        ws.row_dimensions[r].auto_height = False
        ws.row_dimensions[r].height = 2.82 * CM_TO_POINTS

        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)

        b = ws.cell(r, 2)
        b.value = s.get('nameMr') or s.get('name') or ""
        b.alignment = Alignment(horizontal='center', vertical='center')
        b.font = Font(name=font_name, size=14, bold=True)

        thin_range(f'B{r}:K{r}')
        r += 1

    first_row = 3
    last_row = max(2, r - 1)

    for col in range(2, 12):
        edge(first_row, col, top=medium)
        edge(last_row, col, bottom=medium)
    for rr in range(first_row, last_row + 1):
        edge(rr, 2, left=medium)
        edge(rr, 11, right=medium)

    edge(first_row, 2, top=medium); edge(first_row, 3, top=medium)
    edge(last_row, 2, bottom=medium); edge(last_row, 3, bottom=medium)
    for rr in range(first_row, last_row + 1):
        edge(rr, 2, left=medium); edge(rr, 3, right=medium)

    footer_row1 = last_row + 1
    ws.row_dimensions[footer_row1].auto_height = False
    ws.row_dimensions[footer_row1].height = 1.00 * CM_TO_POINTS

    footer_row2 = footer_row1 + 1
    ws.merge_cells(start_row=footer_row2, start_column=10, end_row=footer_row2, end_column=11)
    lab = ws.cell(footer_row2, 10)
    lab.value = "वर्गशिक्षक"
    lab.alignment = Alignment(horizontal='center', vertical='center')
    lab.font = Font(name=font_name, size=14, bold=False)

    footer_row3 = footer_row2 + 1
    ws.row_dimensions[footer_row3].auto_height = False
    ws.row_dimensions[footer_row3].height = 0.90 * CM_TO_POINTS

    def resolve_teacher_name() -> str:
        if isinstance(catalog_doc, dict):
            v = catalog_doc.get('classTeacher')
            if isinstance(v, str) and v.strip():
                return v.strip()
            if isinstance(class_no, str) and isinstance(division, str):
                key = f"catalog-{class_no}-{division}"
                nested = catalog_doc.get(key)
                if isinstance(nested, dict):
                    v2 = nested.get('classTeacher')
                    if isinstance(v2, str) and v2.strip():
                        return v2.strip()
        if isinstance(class_no, dict):
            v = class_no.get('classTeacher')
            if isinstance(v, str) and v.strip():
                return v.strip()
        if isinstance(division, dict):
            v = division.get('classTeacher')
            if isinstance(v, str) and v.strip():
                return v.strip()
        for src in (catalog_doc, class_no, division):
            try:
                v = getattr(src, 'classTeacher', None)
                if isinstance(v, str) and v.strip():
                    return v.strip()
            except Exception:
                pass
        return ""

    teacher_name = resolve_teacher_name()
    if len(teacher_name) >= 2 and teacher_name[0] == '"' and teacher_name[-1] == '"':
        teacher_name = teacher_name[1:-1].strip()

    footer_row4 = footer_row3 + 1
    ws.merge_cells(start_row=footer_row4, start_column=10, end_row=footer_row4, end_column=11)
    ct = ws.cell(footer_row4, 10)
    ct.value = teacher_name
    ct.alignment = Alignment(horizontal='center', vertical='center')
    ct.font = Font(name=font_name, size=14, bold=False)

    return wb
