from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image
from PIL import Image as PILImage


def to_marathi_numerals(number: int) -> str:
    english_digits = '0123456789'
    marathi_digits = '०१२३४५६७८९'
    translation_table = str.maketrans(english_digits, marathi_digits)
    return str(number).translate(translation_table)


def add_front_page_fn(
    wb: Workbook,
    report_data: Dict[str, Any],
    assets_dir: Path | None = None,
) -> Workbook:
    # Helpers
    def apply_thin_grid(ws, cell_range_str: str):
        thin_side = Side(border_style="thin")
        thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
        target = ws[cell_range_str]
        if isinstance(target, Cell):
            target.border = thin_border
        else:
            for row in target:
                for cell in row:
                    cell.border = thin_border

    def apply_medium_box_border(ws, cell_range_str: str):
        medium_side = Side(border_style="medium")
        target = ws[cell_range_str]
        if isinstance(target, Cell):
            target.border = Border(top=medium_side, left=medium_side, right=medium_side, bottom=medium_side)
            return
        rows = list(target)
        for cell in rows[0]:
            existing = cell.border.copy(); existing.top = medium_side; cell.border = existing
        for cell in rows[-1]:
            existing = cell.border.copy(); existing.bottom = medium_side; cell.border = existing
        for row in rows:
            cell = row[0]; existing = cell.border.copy(); existing.left = medium_side; cell.border = existing
        for row in rows:
            cell = row[-1]; existing = cell.border.copy(); existing.right = medium_side; cell.border = existing

    ws = wb.create_sheet("Front Page", 0)

    # Page setup and dimensions
    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
    ws.page_margins.left = 0.0
    ws.page_margins.right = 0.0
    ws.page_margins.top = 0.3937
    ws.page_margins.bottom = 0.3937
    ws.column_dimensions['A'].width = 10.53
    ws.column_dimensions['B'].width = 14.34
    col_width_c_to_j = 5.62
    for i in range(3, 11):
        ws.column_dimensions[get_column_letter(i)].width = col_width_c_to_j
    ws.column_dimensions['K'].width = 15.79
    ws.column_dimensions['L'].width = 5.82
    ws.column_dimensions['M'].width = 5.82
    ws.column_dimensions['N'].width = 5.82
    ws.column_dimensions['O'].width = 4.01
    ws.row_dimensions[1].height = 15.03
    ws.row_dimensions[2].height = 22.11
    ws.row_dimensions[3].height = 28.92
    ws.row_dimensions[4].height = 16.73
    ws.row_dimensions[5].height = 22.11
    ws.row_dimensions[6].height = 19.28
    ws.row_dimensions[7].height = 24.66
    ws.row_dimensions[8].height = 31.19
    ws.row_dimensions[9].height = 23.53
    ws.row_dimensions[10].height = 38.27
    ws.row_dimensions[11].height = 22.11
    for i in range(12, 19):
        ws.row_dimensions[i].height = 24.10
    ws.row_dimensions[19].height = 20.41
    ws.row_dimensions[20].height = 20.41
    ws.row_dimensions[21].height = 26.37
    ws.row_dimensions[22].height = 19.85
    ws.row_dimensions[23].height = 18.71
    for i in range(24, 32):
        ws.row_dimensions[i].height = 16.16
    ws.row_dimensions[32].height = 14.46
    for i in range(33, 36):
        ws.row_dimensions[i].height = 19.28
    ws.row_dimensions[36].height = 25.23
    ws.row_dimensions[37].height = 19.28

    if assets_dir:
        logo_path = assets_dir / "School_logo.png"
        if logo_path.exists():
            try:
                with PILImage.open(logo_path) as pil_img:
                    original_width, original_height = pil_img.size
                img = Image(str(logo_path))
                target_height = 135
                aspect_ratio = original_width / max(1, original_height)
                img.height = target_height
                img.width = int(target_height * aspect_ratio)
                ws.add_image(img, 'D2')
            except Exception:
                pass

    kokila_available = assets_dir and (assets_dir / "Kokila.ttf").exists()
    font_name = "Kokila" if kokila_available else "Calibri"
    header_font_r2 = Font(name=font_name, size=16, bold=True)
    header_font_r3 = Font(name=font_name, size=22, bold=True)
    header_font_r4 = Font(name=font_name, size=12, bold=True)
    header_font_r5 = Font(name=font_name, size=16, bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('F2:N2')
    c = ws['F2']; c.value = "श्री. अगस्ति एज्युकेशन सोसायटी मुंबई संचालित..."; c.font = header_font_r2; c.alignment = center_alignment
    ws.merge_cells('F3:N3')
    c = ws['F3']; c.value = "अगस्ति विद्यालय, अकोले"; c.font = header_font_r3; c.alignment = center_alignment
    ws.merge_cells('F4:N4')
    c = ws['F4']; c.value = "ता. अकोले, जि. अहिल्यानगर"; c.font = header_font_r4; c.alignment = center_alignment
    ws.merge_cells('F5:N5')
    c = ws['F5']; c.value = "महिनावार उपस्थिती व अभ्यासक्रम"; c.font = header_font_r5; c.alignment = center_alignment

    ws.merge_cells('B7:E7')
    ws.merge_cells('L7:N7')
    ws.merge_cells('B8:F8')
    ws.merge_cells('B9:B11')
    ws.merge_cells('C9:N9')
    ws.merge_cells('C10:D10')
    ws.merge_cells('E10:F10')
    ws.merge_cells('G10:H10')
    ws.merge_cells('I10:J10')
    ws.merge_cells('K10:N10')

    header_ranges_to_border = ['B9:B11', 'C9:N9', 'C10:D10', 'E10:F10', 'G10:H10', 'I10:J10', 'K10:N10', 'C11:D11', 'E11:F11', 'G11:H11', 'I11:J11', 'K11', 'L11:N11']
    for r in header_ranges_to_border:
        apply_thin_grid(ws, r)
        apply_medium_box_border(ws, r)

    table_header_font = Font(name=font_name, size=14, bold=True)
    table_subheader_font = Font(name=font_name, size=14, bold=False)
    table_label_font = Font(name=font_name, size=12)
    table_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    table_left = Alignment(horizontal='left', vertical='center')
    table_right = Alignment(horizontal='right', vertical='center')

    c = ws['B9']; c.value = "विद्यार्थ्यांची वर्गवारी"; c.font = table_header_font; c.alignment = table_center
    c = ws['C9']; c.value = "विद्यार्थी संख्या"; c.font = table_header_font; c.alignment = table_center
    for cell_ref, text in [('C10', "महिन्याचा पहिला दिवस"), ('E10', "प्रवेश दिलेले"), ('G10', "नाव कमी / तुकडी बदल"), ('I10', "शेवटच्या दिवशी"), ('K10', "मागासवर्गीय")]:
        c = ws[cell_ref]; c.value = text; c.font = table_subheader_font; c.alignment = table_center
    for cell_ref, text in {'C11': "मुले", 'E11': "मुले", 'G11': "मुले", 'I11': "मुले", 'L11': "मुले", 'D11': "मुली", 'F11': "मुली", 'H11': "मुली", 'J11': "मुली", 'M11': "मुली", 'K11': "प्रवर्ग", 'N11': "एकूण"}.items():
        c = ws[cell_ref]; c.value = text; c.font = table_subheader_font; c.alignment = table_center
    for i, text in enumerate(["मोफत शिक्षण", "बी.सी. एकदा नापास", "प्रा. शिक्षक पाल्य", "मा. शिक्षक पाल्य", "माजी सैनिक", "आजी सैनिक", "दार", "एकूण", "एकूण"], start=12):
        c = ws[f'B{i}']; c.value = text; c.font = table_label_font; c.alignment = table_left
    for i, text in enumerate(["अनु. जाती ( SC )", "अनु. जमाती (ST )", "भ. वि. जा. (NT)", "विशेष मागास (SBC)", "इतर मागास (OBC)", "खुला (OPEN)", "एकूण :-", "पटावर :-", "सरासरी हजेरी :-"], start=12):
        c = ws[f'K{i}']; c.value = text; c.font = table_label_font; c.alignment = table_left

    apply_thin_grid(ws, 'B12:N20')
    ws.merge_cells('L19:N19')
    ws.merge_cells('C20:D20')
    ws.merge_cells('E20:F20')
    ws.merge_cells('G20:H20')
    ws.merge_cells('I20:J20')
    ws.merge_cells('L20:N20')
    for r in ['B12:B18', 'B19', 'B20', 'C12:D18', 'C19:D19', 'C20:D20', 'E12:F18', 'E19:F19', 'E20:F20', 'G12:H18', 'G19:H19', 'G20:H20', 'I12:J18', 'I19:J19', 'I20:J20', 'K12:K18', 'K19', 'K20', 'L12:N18', 'L19:N19', 'L20:N20']:
        apply_medium_box_border(ws, r)

    apply_thin_grid(ws, 'B22:N31')
    ws.merge_cells('B22:N22')
    for i in range(23, 32):
        ws.merge_cells(f'C{i}:I{i}')
    for i in range(23, 32):
        ws.merge_cells(f'J{i}:N{i}')
    for r in ['B22:N22', 'B23:N23', 'B23:B31', 'C23:I31', 'J23:N31']:
        apply_medium_box_border(ws, r)

    curriculum_header_font = Font(name=font_name, size=14, bold=True)
    c = ws['B22']; c.value = "कमी केलेल्या व प्रवेश दिलेल्या विद्यार्थ्यांची नावे"; c.font = curriculum_header_font; c.alignment = table_center
    c = ws['B23']; c.value = "रजि. नंबर"; c.font = curriculum_header_font; c.alignment = table_center
    c = ws['C23']; c.value = "विद्यार्थ्याचे नाव"; c.font = curriculum_header_font; c.alignment = table_center
    c = ws['J23']; c.value = "शेरा"; c.font = curriculum_header_font; c.alignment = table_center

    ws.merge_cells('B33:N33')
    ws.merge_cells('B34:N34')
    ws.merge_cells('L35:N35')
    ws.merge_cells('L37:N37')

    signature_font = Font(name=font_name, size=14)
    c = ws['B33']; c.value = "मी असे प्रमाणित करतो की, वरील नोंदी बरोबर आहेत."; c.font = signature_font; c.alignment = table_left
    c = ws['B34']; c.value = "हजेरीपत्रक तपासले असून त्यातील नोंदी आमचे माहितीनुसार बरोबर आहेत"; c.font = signature_font; c.alignment = table_left
    c = ws['L35']; c.value = "वर्गशिक्षक"; c.font = signature_font; c.alignment = table_center

    # Dynamic data
    teacher_name = report_data.get('teacher_name', 'N/A')
    c = ws['B8']; c.value = f"वर्गशिक्षक :- {teacher_name}"; c.font = Font(name=font_name, size=14, bold=True); c.alignment = table_left
    c = ws['L37']; c.value = teacher_name; c.font = signature_font; c.alignment = table_center

    # Decide month/year: historical override if provided, else current
    marathi_months = ["जानेवारी", "फेब्रुवारी", "मार्च", "एप्रिल", "मे", "जून", "जुलै", "ऑगस्ट", "सप्टेंबर", "ऑक्टोबर", "नोव्हेंबर", "डिसेंबर"]
    sel_month = report_data.get("selected_month")
    sel_year = report_data.get("selected_year")
    if isinstance(sel_month, int) and 1 <= sel_month <= 12 and isinstance(sel_year, int) and sel_year > 0:
        month_num = sel_month
        year = sel_year
    else:
        today = datetime.today()
        month_num = today.month
        year = today.year
    month_name = marathi_months[month_num - 1]
    marathi_year = to_marathi_numerals(year)

    info_font = Font(name=font_name, size=18, bold=True)
    c = ws['B7']; c.value = f"महिना: {month_name} {marathi_year}"; c.font = info_font; c.alignment = table_left
    class_name = report_data.get('class_name_mr', '')
    c = ws['K7']; c.value = f"इयत्ता: {class_name}"; c.font = info_font; c.alignment = table_right

    # Division label
    marathi_division_map = {'A': 'अ', 'B': 'ब', 'C': 'क', 'D': 'ड', 'E': 'ई', 'F': 'फ'}
    raw_value = report_data.get('division_name_mr') or report_data.get('division', '')
    cleaned_value = str(raw_value).strip().upper()
    final_division_name = marathi_division_map.get(cleaned_value, str(raw_value).strip())
    c = ws['L7']; c.value = f"तुकडी: {final_division_name}"; c.font = info_font; c.alignment = table_center

    return wb
