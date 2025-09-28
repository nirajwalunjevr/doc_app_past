from __future__ import annotations
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_catalog_excel_fn(
    class_no: str | int,
    division: str,
    students: List[Dict[str, Any]],
    last_girl_row_hint: Optional[int] = None,
) -> Workbook:
    class_division_str = f"{class_no}-{division.upper()}"

    last_girl_row = last_girl_row_hint
    if last_girl_row is None:
        for idx, st in enumerate(students):
            if st.get('gender') == 'मुलगी':
                last_girl_row = idx + 3  # data starts at row 3

    wb = Workbook()
    ws = wb.active
    ws.title = f"Catalog Class {class_division_str}"

    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
    ws.page_margins.left = 0.0
    ws.page_margins.right = 0.0
    ws.page_margins.top = 0.3937
    ws.page_margins.bottom = 0.3937

    ws.insert_cols(idx=1, amount=1)       # A
    ws.insert_cols(idx=11, amount=2)      # K, L
    ws.insert_cols(idx=46, amount=1)      # AT

    header_font = Font(name='Kokila', size=14, bold=True)
    body_font = Font(name='Kokila', size=14, bold=False)
    small_header_font = Font(name='Kokila', size=10, bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(border_style='thin'), right=Side(border_style='thin'),
        top=Side(border_style='thin'), bottom=Side(border_style='thin')
    )
    medium_side = Side(border_style='medium')

    ws.merge_cells('B1:J1')
    roster_headers = ['रजि. नं.', 'सवलत', 'जात', 'प्रवर्ग', 'Category',
                      'जन्म दिनांक', 'अ.न.', 'विद्यार्थ्यांचे नाव', 'आईचे नाव']
    for i, text in enumerate(roster_headers):
        cidx = 2 + i  # B..J
        cell = ws.cell(row=2, column=cidx)
        cell.value = text
        cell.font = header_font
        cell.alignment = center_align

    ws.merge_cells('M1:AJ1')
    ws.merge_cells('AK1:AQ1')
    c = ws['AK1']; c.value = 'एकूण दिवस'; c.font = header_font; c.alignment = center_align

    day_headers = [str(i) for i in range(1, 32)]
    extra_headers = ['कामाचे दिवस', 'शेरा']
    for i, text in enumerate(day_headers + extra_headers):
        cidx = 13 + i  # M..AQ
        cell = ws.cell(row=2, column=cidx)
        cell.value = text
        cell.font = small_header_font
        cell.alignment = center_align

    ws['F2'].font = Font(name='Kokila', size=9, bold=True)
    ws['AR2'].font = Font(name='Kokila', size=8, bold=True)
    ws['AS2'].font = Font(name='Kokila', size=12, bold=True)

    for r, st in enumerate(students, start=3):
        dob = st.get('dob')
        if hasattr(dob, 'strftime'):
            dob_str = dob.strftime('%d-%m-%Y')
        else:
            dob_str = dob or ''
        row_vals = [
            st.get('regNo', ''), st.get('concession', ''), st.get('caste', ''),
            st.get('categoryMr', ''), st.get('categoryEn', ''), dob_str,
            st.get('rollNo', ''), st.get('fullNameMr', ''), st.get('motherName', '')
        ]
        for off, val in enumerate(row_vals):
            c = ws.cell(row=r, column=2 + off)
            c.value = val
            c.font = body_font
            c.alignment = center_align if (2 + off) in {2, 8} else left_align

    for col_letter, cm in {
        'B': 1.48, 'C': 1.70, 'D': 2.31, 'E': 1.67, 'F': 1.60,
        'G': 2.50, 'H': 1.16, 'I': 4.80, 'J': 2.31
    }.items():
        ws.column_dimensions[col_letter].width = cm / 0.2117
    for i in range(13, 44):  # M..AQ
        ws.column_dimensions[get_column_letter(i)].width = 0.60 / 0.2117
    ws.column_dimensions['AR'].width = (33 - 5) / 7.0
    ws.column_dimensions['AS'].width = (30 - 5) / 7.0
    ws.column_dimensions['AT'].width = (45 - 5) / 7.0
    ws.column_dimensions['A'].width  = (45 - 5) / 7.0
    ws.column_dimensions['K'].width  = (75.6 - 5) / 7.0
    ws.column_dimensions['L'].width  = (75.6 - 5) / 7.0

    ws.row_dimensions[18].auto_height = False
    ws.row_dimensions[8].auto_height = False
    ws.row_dimensions[18].height = 22.68
    ws.row_dimensions[8].height = 28.35

    target_pt = 0.87 * 28.3465
    last_student_row = 2 + len(students)
    for r in range(3, last_student_row + 1):
        ws.row_dimensions[r].auto_height = False
        ws.row_dimensions[r].height = target_pt

    border_end_row = last_student_row + 4
    footer_start_row = border_end_row + 1
    footer_rows = [footer_start_row, footer_start_row + 1, footer_start_row + 2]

    for r in range(last_student_row + 1, border_end_row + 1):
        ws.row_dimensions[r].auto_height = False
        ws.row_dimensions[r].height = target_pt
    for r in footer_rows:
        ws.row_dimensions[r].auto_height = False
        ws.row_dimensions[r].height = target_pt

    for row in ws.iter_rows(min_row=1, max_row=border_end_row, min_col=2, max_col=10):
        for c in row:
            c.border = Border(
                left=Side(border_style='thin'), right=Side(border_style='thin'),
                top=Side(border_style='thin'), bottom=Side(border_style='thin')
            )
    for row in ws.iter_rows(min_row=1, max_row=border_end_row, min_col=13, max_col=45):
        for c in row:
            c.border = Border(
                left=Side(border_style='thin'), right=Side(border_style='thin'),
                top=Side(border_style='thin'), bottom=Side(border_style='thin')
            )
    for row in ws.iter_rows(min_row=1, max_row=border_end_row, min_col=37, max_col=43):
        for c in row:
            c.border = Border(
                left=Side(border_style='thin'), right=Side(border_style='thin'),
                top=Side(border_style='thin'), bottom=Side(border_style='thin')
            )

    label_font = Font(name='Kokila', size=14, bold=True)
    mid_center = Alignment(horizontal='center', vertical='center')
    for r, txt in zip(footer_rows, ['हजर', 'गैरहजर', 'एकूण']):
        cell = ws.cell(row=r, column=9)  # I
        cell.value = txt
        cell.font = label_font
        cell.alignment = mid_center

    for r in footer_rows:
        for row in ws.iter_rows(min_row=r, max_row=r, min_col=2, max_col=10):
            for c in row:
                c.border = Border(
                    left=Side(border_style='thin'), right=Side(border_style='thin'),
                    top=Side(border_style='thin'), bottom=Side(border_style='thin')
                )
        for row in ws.iter_rows(min_row=r, max_row=r, min_col=13, max_col=45):
            for c in row:
                c.border = Border(
                    left=Side(border_style='thin'), right=Side(border_style='thin'),
                    top=Side(border_style='thin'), bottom=Side(border_style='thin')
                )
        for row in ws.iter_rows(min_row=r, max_row=r, min_col=37, max_col=43):
            for c in row:
                c.border = Border(
                    left=Side(border_style='thin'), right=Side(border_style='thin'),
                    top=Side(border_style='thin'), bottom=Side(border_style='thin')
                )

    def outline_row_span(row, c_start, c_end):
        for cidx in range(c_start, c_end + 1):
            ws.cell(row=row, column=cidx).border = Border(
                top=Side(border_style='medium'), bottom=Side(border_style='medium'),
                left=Side(border_style='medium') if cidx == c_start else Side(border_style='thin'),
                right=Side(border_style='medium') if cidx == c_end else Side(border_style='thin')
            )

    def outline_merged_span_edges(row_top, row_bottom, col_left, col_right):
        medium = Side(border_style='medium')
        for cidx in range(col_left, col_right + 1):
            existing = ws.cell(row=row_top, column=cidx).border
            ws.cell(row=row_top, column=cidx).border = Border(
                top=medium,
                left=medium if cidx == col_left else existing.left,
                right=medium if cidx == col_right else existing.right,
                bottom=existing.bottom
            )
        for cidx in range(col_left, col_right + 1):
            existing = ws.cell(row=row_bottom, column=cidx).border
            ws.cell(row=row_bottom, column=cidx).border = Border(
                bottom=medium,
                top=existing.top,
                left=medium if cidx == col_left else existing.left,
                right=medium if cidx == col_right else existing.right
            )
        for r in range(row_top, row_bottom + 1):
            existing = ws.cell(row=r, column=col_left).border
            ws.cell(row=r, column=col_left).border = Border(
                left=medium, top=existing.top, bottom=existing.bottom, right=existing.right
            )
            existing = ws.cell(row=r, column=col_right).border
            ws.cell(row=r, column=col_right).border = Border(
                right=medium, top=existing.top, bottom=existing.bottom, left=existing.left
            )

    def outline_cell(r, cidx):
        medium = Side(border_style='medium')
        ws.cell(row=r, column=cidx).border = Border(
            top=medium, bottom=medium, left=medium, right=medium
        )

    outline_merged_span_edges(1, 1, 2, 10)      # B1:J1
    outline_row_span(2, 2, 10)                 # B2:J2
    outline_merged_span_edges(1, 1, 13, 45)     # M1:AS1
    outline_row_span(2, 13, 43)                 # M2:AQ2
    ws.merge_cells("AR1:AS1")
    outline_cell(1, 37)                          # AK1
    outline_cell(1, 44)                          # AR1
    outline_cell(1, 45)                          # AS1
    outline_cell(2, 44)                          # AR2
    outline_cell(2, 45)                          # AS2

    printable_height_pt = (35.56 - 2.0) * 28.3465
    total_height = 0
    page_boundaries: list[int] = []

    def get_row_height(r):
        h = ws.row_dimensions[r].height
        return h if h else ws.row_dimensions[0].height or 15

    for r in range(3, footer_rows[0]):
        total_height += get_row_height(r)
        if total_height > printable_height_pt:
            page_boundaries.append(r - 3)
            total_height = get_row_height(r)
    page_boundaries.append(footer_rows[0] - 1)

    start_row = 3
    for end_row in page_boundaries:
        outline_merged_span_edges(start_row, end_row, 2, 10)
        outline_merged_span_edges(start_row, end_row, 13, 45)
        start_row = end_row + 1

    outline_merged_span_edges(footer_start_row, footer_start_row + 2, 2, 10)
    outline_merged_span_edges(footer_start_row, footer_start_row + 2, 13, 45)

    last_footer_row = footer_rows[-1]

    def outline_col_span(col_idx, r_start, r_end):
        medium = Side(border_style='medium')
        for r in range(r_start, r_end + 1):
            cell = ws.cell(row=r, column=col_idx)
            existing = cell.border
            top_side = existing.top if getattr(existing.top, 'style', None) == 'medium' else (medium if r == r_start else Side(border_style='thin'))
            bottom_side = existing.bottom if getattr(existing.bottom, 'style', None) == 'medium' else (medium if r == r_end else Side(border_style='thin'))
            left_side = existing.left if getattr(existing.left, 'style', None) == 'medium' else medium
            right_side = existing.right if getattr(existing.right, 'style', None) == 'medium' else medium
            cell.border = Border(left=left_side, right=right_side, top=top_side, bottom=bottom_side)

    page_boundaries_with_footer = page_boundaries[:-1] + [last_footer_row]
    start_row = 1
    for end_row in page_boundaries_with_footer:
        outline_col_span(44, start_row, end_row)  # AR
        outline_col_span(45, start_row, end_row)  # AS
        start_row = end_row + 1

    for col_idx in (44, 45):
        for r in (1, 2):
            medium = Side(border_style='medium')
            ws.cell(row=r, column=col_idx).border = Border(
                top=medium, bottom=medium, left=medium, right=medium
            )

    if last_girl_row:
        for col_idx in range(2, 11):
            cell = ws.cell(row=last_girl_row, column=col_idx)
            cell.border = Border(
                left=cell.border.left, right=cell.border.right,
                top=cell.border.top, bottom=Side(border_style='medium')
            )
        for col_idx in range(13, 46):
            cell = ws.cell(row=last_girl_row, column=col_idx)
            cell.border = Border(
                left=cell.border.left, right=cell.border.right,
                top=cell.border.top, bottom=Side(border_style='medium')
            )

    ws.row_dimensions[1].auto_height = False
    ws.row_dimensions[1].height = 25.35
    _ = ws['B1'].alignment

    return wb
